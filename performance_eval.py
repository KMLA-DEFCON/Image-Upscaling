import cv2, os, shutil
from skimage.metrics import peak_signal_noise_ratio, structural_similarity
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color

wb = load_workbook(filename = "./result.xlsx")
ws = wb.active

image_scale = 0
error_list_psnr = []
error_list_ssim = []

def read_image(path):

    img = cv2.imread(path)    
    size = img.shape
    dimension = (size[0], size[1])

    return img, size, dimension

def image_change_scale(img, dimension, scale=100, interpolation=cv2.INTER_LINEAR):
    scale /= 100
    new_dimension = (int(dimension[1]*scale), int(dimension[0]*scale))
    resized_img = cv2.resize(img, new_dimension, interpolation=interpolation)
    return resized_img

def get_error(image_file):
    global image_scale, error_list_psnr, error_list_ssim
    images_list = {}
    error_list_psnr = []
    error_list_ssim = []

    # Read Image
    img, size, dimension = read_image(image_file)
    images_list['Original Image'] = img

    # Change Image Size
    scale_percent = 20 # percent of original image size
    image_scale = scale_percent
    resized_img = image_change_scale(img, dimension, scale_percent)
    images_list['Smalled Image'] = resized_img

    # Change image to original size using nearest neighbor interpolation
    nn_img = image_change_scale(
        resized_img, dimension, interpolation=cv2.INTER_NEAREST)
    images_list['Nearest Neighbor Interpolation'] = nn_img

    # Change image to original size using bilinear interpolation
    bil_img = image_change_scale(
        resized_img, dimension, interpolation=cv2.INTER_LINEAR)
    images_list['Bilinear Interpolation'] = bil_img

    # Change image to original size using cubiclinear interpolation (4*4 pixel neighborhood)
    cubic_img = image_change_scale(
        resized_img, dimension, interpolation=cv2.INTER_CUBIC)
    images_list['CubicLinear Interpolation'] = cubic_img

    # Change image to original size using lanczos interpolation (8*8 pixel neighborhood)
    czos_img = image_change_scale(
        resized_img, dimension, interpolation=cv2.INTER_LANCZOS4)
    images_list['Lanczos Interpolation'] = czos_img

    # error calculate between the smalled image and the original image
    
    error_list_psnr.append(peak_signal_noise_ratio(nn_img, img))
    error_list_psnr.append(peak_signal_noise_ratio(bil_img, img))
    error_list_psnr.append(peak_signal_noise_ratio(cubic_img, img))
    error_list_psnr.append(peak_signal_noise_ratio(czos_img, img))

    error_list_ssim.append(structural_similarity(nn_img, img, channel_axis=2))
    error_list_ssim.append(structural_similarity(bil_img, img, channel_axis=2))
    error_list_ssim.append(structural_similarity(cubic_img, img, channel_axis=2))
    error_list_ssim.append(structural_similarity(czos_img, img, channel_axis=2)) 

interpolation_methods = ["Nearest Neighbor", "Bilinear", "Cubiclinear", "Lanczos"]
num = 2
color_index = 40


for image_type in os.listdir("./images"):

    ascii_c = 3
    while True:
        if ws.cell(row=num, column=ascii_c).value is None:
            break
        
        else:
            ascii_c +=1

    for image in os.listdir(f"./images/{image_type}"):

        get_error(f"./images/{image_type}/{image}")

        ws.cell(row=num, column=ascii_c).value = str(image)
        color = Color(indexed=color_index) 
        paint_cell = PatternFill(patternType='solid', fgColor=color)
        ws.cell(row=num, column=ascii_c).fill = paint_cell
        ws.cell(row=num-1, column=ascii_c).value = f"{image_type} / {image_scale}"

        for x in range(4):
            ws.cell(row=num+1+x, column=ascii_c).value = error_list_psnr[x]

            if error_list_psnr[x] == max(error_list_psnr):
                ws.cell(row=num+1+x, column=ascii_c).fill = PatternFill(patternType='solid', fgColor=Color(indexed=3))
            
            elif error_list_psnr[x] == min(error_list_psnr):
                ws.cell(row=num+1+x, column = ascii_c).fill = PatternFill(patternType='solid', fgColor=Color(indexed=2))

            ws.cell(row=num+6+x, column=ascii_c).value = error_list_ssim[x]

            if error_list_ssim[x] == max(error_list_ssim):
                ws.cell(row=num+6+x, column=ascii_c).fill = PatternFill(patternType='solid', fgColor=Color(indexed=3))
            
            elif error_list_ssim[x] == min(error_list_ssim):
                ws.cell(row=num+6+x, column = ascii_c).fill = PatternFill(patternType='solid', fgColor=Color(indexed=2))
            
        shutil.move(f"./images/{image_type}/{image}", f"./saved_images/{image_type}/{image}")

        ascii_c+=1

    color_index+=1
    num += 13
        
wb.save("result.xlsx")
